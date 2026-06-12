from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = (
    ROOT
    / "matriz_consistencia"
    / "Matriz_de_Consistencia_Tesis_Cables_Heterogeneidad_Termica.xlsx"
)


def main() -> None:
    temporary = WORKBOOK.with_name(f"{WORKBOOK.stem}.tmp.xlsx")
    workbook = load_workbook(WORKBOOK)

    matrix = workbook["Matriz UNI"]
    matrix["B5"] = (
        "PROBLEMA GENERAL\n"
        "¿Cuál es el efecto de la heterogeneidad térmica espacial del entorno, "
        "representada mediante k(x,y), sobre la temperatura máxima del conductor y "
        "la ampacidad de cables eléctricos subterráneos en escenarios representativos "
        "de centrales solares y eólicas, según un modelo numérico 2D previamente "
        "verificado y validado?"
    )
    matrix["C5"] = (
        "OBJETIVO GENERAL\n"
        "Determinar y cuantificar el efecto de la heterogeneidad térmica espacial del "
        "suelo, bedding y backfill sobre la temperatura máxima del conductor y la "
        "ampacidad de cables eléctricos subterráneos en centrales solares y eólicas, "
        "mediante la aplicación de un modelo numérico 2D previamente verificado y "
        "validado."
    )
    matrix["F5"] = (
        "Enfoque: cuantitativo.\n"
        "Tipo: aplicada/tecnológica.\n"
        "Nivel: explicativo-predictivo.\n"
        "Diseño: experimental numérico in silico, controlado, factorial y de "
        "sensibilidad; régimen 2D estacionario como núcleo y casos transitorios "
        "seleccionados si el alcance lo permite.\n"
        "Métodos: hipotético-deductivo, modelado numérico y Design Science Research "
        "para construir/evaluar el artefacto computacional.\n"
        "Unidad de análisis: sección transversal 2D de una instalación de cables "
        "subterráneos.\n"
        "Población: posibles secciones 2D de circuitos subterráneos en centrales "
        "renovables. Muestra: no probabilística e intencional, formada por casos de "
        "referencia y escenarios factoriales basados en IEC 60287/60853, Aras (2005), "
        "Kim et al. (2025) y configuraciones representativas.\n"
        "Técnicas/instrumentos: revisión documental; matriz paramétrica; solver PINN "
        "2D en Python; referencia FEM/analítica; datos publicados; archivos CSV y "
        "registros reproducibles.\n"
        "Control de calidad: verificación numérica frente a soluciones analíticas o "
        "FEM con convergencia demostrada; validación física frente a datos "
        "experimentales o publicados cuando estén disponibles. IEC se usa como línea "
        "base normativa, no como verdad de referencia para campos heterogéneos.\n"
        "Análisis: error de Tmax y T(x,y), residuos PDE/BC, balance de energía, "
        "incertidumbre del modelo, ΔI%, curvas de respuesta, regresión/sensibilidad y "
        "repetición con semillas. La incertidumbre numérica debe ser menor que el "
        "efecto térmico que se pretende atribuir a k(x,y)."
    )
    matrix["B6"] = (
        "PE1. ¿Cómo influyen la magnitud y el contraste de la conductividad térmica "
        "k(x,y) en la temperatura máxima del conductor y el campo térmico cuando se "
        "mantienen constantes la geometría, la carga y la disposición espacial?"
    )
    matrix["C6"] = (
        "OE1. Determinar el efecto de la magnitud y el contraste de k(x,y) sobre Tmax "
        "y T(x,y) mediante comparaciones controladas con un entorno homogéneo "
        "equivalente."
    )
    matrix["D6"] = (
        "HE1. La disminución de k en el entorno y el aumento del contraste térmico "
        "incrementarán Tmax y los gradientes térmicos respecto del caso homogéneo "
        "equivalente."
    )
    matrix["E6"] = (
        "X: kmin, kmedia, Ck y CVk.\n"
        "Y1: Tmax, ΔTmax, T(x,y) y gradiente térmico.\n"
        "Controles: geometría, carga, patrón espacial, profundidad y BC."
    )
    matrix.row_dimensions[5].height = max(
        matrix.row_dimensions[5].height or 0, 158
    )

    activities = workbook["Matriz por actividades"]
    activity_values = [
        "3. Verificación numérica y validación física",
        "Necesidad metodológica: antes de atribuir cambios de Tmax a k(x,y), debe "
        "demostrarse que la discrepancia e incertidumbre del modelo son menores que "
        "el efecto térmico analizado.",
        "Verificar que el solver resuelva correctamente la ecuación de calor y validar "
        "que represente la respuesta física dentro de un dominio de aplicación "
        "explícito.",
        "Informe V&V: pruebas analíticas/FEM, contraste con datos experimentales o "
        "publicados, presupuesto de error y dominio de validez.",
        "No es una variable causal de la investigación; es un control metodológico.",
        "Verificación: solución analítica o manufacturada y FEM con convergencia. "
        "Validación: datos experimentales/publicados. IEC: comparación normativa.",
        "Discrepancia e incertidumbre del modelo.",
        "No corresponde una hipótesis científica adicional; se aplica un criterio de "
        "aceptación previo a los experimentos de heterogeneidad.",
        "Error de Tmax y T(x,y); residuos PDE/BC; balance de energía; convergencia; "
        "sesgo; dispersión entre semillas; Umodelo/|ΔTefecto|.",
        "Mismos materiales, pérdidas, geometría, dominio y BC; referencia "
        "independiente; malla FEM convergente; semillas registradas.",
    ]
    for column, value in enumerate(activity_values, 1):
        activities.cell(7, column).value = value
    activities.row_dimensions[7].height = max(
        activities.row_dimensions[7].height or 0, 128
    )

    operationalization = workbook[workbook.sheetnames[2]]
    for row in operationalization.iter_rows(min_row=5):
        if row[0].value != "Exactitud del modelo":
            continue
        values = [
            "Exactitud del modelo",
            "Control metodológico",
            "Grado de concordancia entre el modelo propuesto y una referencia "
            "aceptada.",
            "Verificación del solver frente a soluciones conocidas y validación "
            "frente a evidencia física independiente.",
            "Verificación y validación",
            "Error de Tmax/campo; residuos; balance; sesgo; Umodelo/|ΔTefecto|",
            "K; %; adimensional",
            "Solución analítica/manufacturada; FEM convergente; datos "
            "experimentales/publicados",
            "Aceptar solo si hay convergencia y consistencia física, y si la "
            "incertidumbre del modelo es menor que el efecto de heterogeneidad "
            "evaluado; documentar el dominio de validez.",
        ]
        for column, value in enumerate(values, 1):
            operationalization.cell(row[0].row, column).value = value
        operationalization.row_dimensions[row[0].row].height = max(
            operationalization.row_dimensions[row[0].row].height or 0, 75
        )
        break
    else:
        raise RuntimeError("No se encontró la fila 'Exactitud del modelo'.")

    workbook.save(temporary)
    os.replace(temporary, WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
