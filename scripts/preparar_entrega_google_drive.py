from __future__ import annotations

import json
import re
import shutil
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PLAN_TEX = ROOT / "Plan" / "plan_tesis_cables_pinn.tex"
PLAN_AUX = ROOT / "Plan" / "plan_tesis_cables_pinn.aux"
PLAN_BIB = ROOT / "Plan" / "referencias_plan_tesis.bib"
ZOTERO_JSON = ROOT / "zotero_library.json"
ZOTERO_SQLITE = Path.home() / "Zotero" / "zotero.sqlite"

BASE_OUT = ROOT / "Entrega_Google_Drive"

PDF_SEARCH_DIRS = [
    ROOT / "zotero_offline" / "pdfs",
    ROOT / "Estado del arte",
    ROOT / "benchmark_papers",
    ROOT / "Plan",
]

RESEARCH_TYPES = {"article", "incollection", "inproceedings", "conference"}

STANDARD_KEYS = {
    "iec60287",
    "iec60853",
    "iec60502_2_2014",
    "ieee442",
    "cigre2022",
    "cigre2025",
    "epriUndergroundFailures",
    "coes2025ev1874",
    "coes2026ev1948",
    "osinergmin2024pi6",
}

MANUAL_TYPES = {"book", "report", "techreport", "manual", "online", "misc"}

APORTES = {
    "neher1957": "Base clasica para relacionar temperatura, perdidas y capacidad de carga de cables; sirve como referencia para comparar el enfoque PINN con el calculo termico tradicional.",
    "enescu2020": "Resume factores termicos que gobiernan el current rating de cables y ayuda a justificar las variables fisicas del artefacto.",
    "enescu2021": "Sistematiza metodos de rating dinamico y efectos del suelo; sustenta la necesidad de representar el entorno termico de forma explicita.",
    "xing2023": "Antecedente directo de PINN aplicada a conduccion de calor anisotropica, util para formular el campo termico como salida continua.",
    "iec60287": "Linea base normativa para ampacidad en regimen permanente y calculo de perdidas; funciona como comparacion de referencia.",
    "iec60853": "Referencia normativa para rating ciclico/emergencia y secado parcial del suelo; delimita la extension transitoria del problema.",
    "cigre2025": "Guia tecnica para FEM en cable rating; aporta criterios de verificacion, malla, calidad y comparacion numerica.",
    "patankar1980numerical": "Fundamento de transferencia de calor numerica y conservacion; apoya controles de balance energetico y verificacion.",
    "aldulaimi2024": "Ejemplo reciente de FEM y red neuronal para temperatura de cables con suelos y backfill variables; identifica oportunidad frente a dependencia de datos FEM.",
    "oclon2015": "Muestra que el bedding modifica la respuesta termica y puede optimizarse; inspira escenarios de relleno y heterogeneidad.",
    "atoccsa2024": "Caso de optimizacion de ampacidad con backfill termico; aporta configuraciones y sensibilidad de materiales.",
    "raissi2019": "Marco original de PINN para resolver problemas directos e inversos mediante residuos fisicos en la funcion de perdida.",
    "pan2025": "Aplicacion PINN a reconstruccion de campo de temperatura 2D; antecedente metodologico cercano al objetivo de la tesis.",
    "billah2023": "Antecedente de PINN para problemas inversos de transferencia de calor, util para futuras extensiones de identificacion de parametros.",
    "chen2024": "Ejemplo de PINN para conduccion no lineal/transitoria; ayuda a delimitar riesgos si se extiende el alcance estacionario.",
    "liu2026": "Antecedente de PINN para problemas inversos estacionarios de conduccion; refuerza la viabilidad de estimar propiedades termicas.",
    "cuomo2022": "Revision de scientific machine learning y PINN; aporta limitaciones, tendencias y criterios de uso responsable.",
    "lawal2022": "Revision sistematica de PINN que sustenta riesgos de entrenamiento, ponderacion de perdidas y evaluacion.",
    "ren2025": "Revision reciente de evolucion y fundamentos de PINN; permite discutir limitaciones metodologicas y fronteras del enfoque.",
    "peffers2007": "Define la secuencia DSRM usada para estructurar problema, diseno, demostracion, evaluacion y comunicacion.",
    "gregor2013": "Aporta criterios para posicionar contribuciones DSR, utilidad, calidad, eficacia y conocimiento transferible.",
    "khumalo2025": "Evidencia cuantitativa del impacto del secado del suelo en resistividad y ampacidad; justifica escenarios criticos.",
    "kim2025": "Benchmark central con bedding/PAC, temperatura ambiente y suelo; aporta datos para validar geometria y respuesta termica.",
    "epriUndergroundFailures": "Contexto operacional de fallas en sistemas subterraneos; refuerza la relevancia practica del problema.",
    "coes2025ev1874": "Evidencia local peruana de evento asociado a cable subterraneo; conecta la tesis con el objeto de estudio nacional.",
    "coes2026ev1948": "Evidencia local de falla en infraestructura subterranea/terminales; apoya la motivacion aplicada.",
    "osinergmin2024pi6": "Fuente tecnica local sobre restricciones y casos de transmision; contribuye al contexto peruano de cables subterraneos.",
    "fariz2026": "Mapa sistematico de dynamic thermal rating en cables subterraneos; ubica la brecha y retos actuales.",
    "cigre2022": "Casos de calculo para verificar herramientas de cable rating; fuente para benchmarks y validacion cruzada.",
    "delport2024": "Guia metodologica reciente para DSR; apoya la evaluacion formativa y refinamiento del artefacto.",
    "aras2005": "Benchmark clasico de metodos de ampacidad y FEM en cables subterraneos; sirve para comparaciones numericas.",
    "ieee442": "Estandar para medir resistividad termica de suelos y backfill; justifica parametros de entrada y calidad de datos.",
    "shukla2021": "Antecedente de PINN por descomposicion de dominio; opcion metodologica si aparecen discontinuidades fuertes.",
    "anders2005": "Manual especializado de rating en entornos termicos desfavorables; base de conceptos de ampacidad y derating.",
    "carslaw1959conduction": "Fundamento teorico de conduccion de calor; respalda formulacion de PDE y condiciones de frontera.",
    "kakac2018heatConduction": "Referencia moderna de conduccion de calor; ayuda a definir el modelo fisico y casos analiticos.",
    "hahn2012heatConduction": "Referencia teorica para ecuacion de calor y soluciones de conduccion; apoya verificaciones analiticas.",
    "farouki1981thermal": "Monografia sobre propiedades termicas de suelos; sustenta rangos de conductividad/resistividad.",
    "devries1966thermal": "Fuente clasica sobre propiedades termicas de suelos; respalda dependencia con composicion y humedad.",
    "iec60502_2_2014": "Norma de cables XLPE de media tension; aporta limites y configuracion fisica de cables usados en escenarios.",
}

PDF_HINTS = {
    "aldulaimi2024": "Adaptive FEM-BPNN model for predicting underground",
    "anders2005": "Rating of electric power cables in unfavorable thermal environment",
    "aras2005": "Aras_ampacitiy_2005",
    "atoccsa2024": "Optimization of ampacity in high-voltage underground cables",
    "carslaw1959conduction": "conduction-of-heat-in-solids",
    "cigre2022": "Power cable rating examples for calculation tool verification",
    "cigre2025": "Finite element analysis for cable rating calculations",
    "coes2025ev1874": "EV-ITF-2025-1874",
    "coes2026ev1948": "EV-ITF-2026-1948",
    "cuomo2022": "scientific_ml_through_pinns",
    "delport2024": "Delport et al. - 2024",
    "enescu2020": "Thermal assessment of power cables",
    "enescu2021": "Concepts and methods to assess the dynamic thermal",
    "fariz2026": "A systematic mapping of dynamic thermal rating",
    "farouki1981thermal": "Thermal properties of soils",
    "gregor2013": "GregorHevnerMISQ2013",
    "hahn2012heatConduction": "Heat_Conduction__Third_Edition",
    "iec60287": "IEC 60287-1-12023",
    "iec60853": "IEC 60853-32002",
    "iec60502_2_2014": "IEC_60502-2-2014",
    "ieee442": "IEEE guide for thermal resistivity",
    "kakac2018heatConduction": "Heat_Conduction__Fifth_Edition",
    "khumalo2025": "Khumalo_Critical_2025",
    "kim2025": "Kim_2024",
    "lawal2022": "Physics-informed neural network (PINN) evolution",
    "liu2026": "Physics-informed neural networks for solving steady-state",
    "neher1957": "The calculation of the temperature rise and load capability",
    "oclon2015": "Oclon_cable_bedding_2005",
    "pan2025": "Research on the reconstruction of the temperature field",
    "patankar1980numerical": "Numerical_Heat_Transfer_and_Fluid_Flow",
    "peffers2007": "Peffers et al. - 2007",
    "raissi2019": "Physics-informed neural networks A deep learning framework",
    "ren2025": "Physics-informed neural networks a review of methodological",
    "shukla2021": "Parallel physics-informed neural networks",
    "xing2023": "Deep learning method based on physics-informed neural network",
}


def unique_output_dir() -> Path:
    return BASE_OUT


def normalize(text: str) -> str:
    text = text or ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_filename(title: str, ext: str = ".pdf", limit: int = 150) -> str:
    ascii_title = unicodedata.normalize("NFKD", title)
    ascii_title = "".join(ch for ch in ascii_title if not unicodedata.combining(ch))
    ascii_title = re.sub(r"[\\/:*?\"<>|]+", " ", ascii_title)
    ascii_title = re.sub(r"\s+", " ", ascii_title).strip(" .")
    ascii_title = ascii_title[:limit].rstrip(" .")
    return (ascii_title or "documento") + ext


def latex_to_text(text: str) -> str:
    text = text or ""
    replacements = {
        r"\'{a}": "a",
        r"\'{e}": "e",
        r"\'{i}": "i",
        r"\'{o}": "o",
        r"\'{u}": "u",
        r"\'a": "a",
        r"\'e": "e",
        r"\'i": "i",
        r"\'o": "o",
        r"\'u": "u",
        r"\"{o}": "o",
        r"\"o": "o",
        r"\c{s}": "s",
        r"\l{}": "l",
        r"\%": "%",
        r"\&": "&",
        r"\_": "_",
        r"\textendash": "-",
        r"\textemdash": "-",
        r"\kx": "k(x,y)",
        r"\Tmax": "Tmax",
        r"\Imax": "Imax",
        r"\pct": "%",
        r"\mathrm": "",
        r"\small": "",
        r"\scriptsize": "",
        r"\RaggedRight": "",
        r"\arraybackslash": "",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    for command in ["textbf", "textit", "emph", "si", "SI"]:
        text = re.sub(r"\\" + command + r"\{([^{}]*)\}", r"\1", text)
    text = re.sub(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?", "", text)
    text = text.replace("{", "").replace("}", "")
    text = text.replace("$", "")
    text = text.replace("~", " ")
    text = text.replace(r"\,", " ")
    text = text.replace("--", "-")
    return re.sub(r"\s+", " ", text).strip()


def parse_bib(path: Path) -> dict[str, dict[str, str]]:
    text = path.read_text(encoding="utf-8")
    entries: dict[str, dict[str, str]] = {}
    pos = 0
    while True:
        match = re.search(r"@(\w+)\s*\{\s*([^,\s]+)\s*,", text[pos:])
        if not match:
            break
        entry_type, key = match.group(1).lower(), match.group(2)
        start = pos + match.end()
        brace_level = 1
        i = start
        while i < len(text) and brace_level:
            if text[i] == "{":
                brace_level += 1
            elif text[i] == "}":
                brace_level -= 1
            i += 1
        body = text[start : i - 1]
        entries[key] = {"ENTRYTYPE": entry_type, "ID": key}
        entries[key].update(parse_bib_fields(body))
        pos = i
    return entries


def parse_bib_fields(body: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    pos = 0
    while pos < len(body):
        match = re.search(r"([A-Za-z][A-Za-z0-9_]+)\s*=", body[pos:])
        if not match:
            break
        name = match.group(1).lower()
        value_start = pos + match.end()
        while value_start < len(body) and body[value_start].isspace():
            value_start += 1
        if value_start >= len(body):
            break
        quote = body[value_start]
        if quote in "{":
            brace_level = 1
            i = value_start + 1
            while i < len(body) and brace_level:
                if body[i] == "{":
                    brace_level += 1
                elif body[i] == "}":
                    brace_level -= 1
                i += 1
            value = body[value_start + 1 : i - 1]
            pos = i
        elif quote == '"':
            i = value_start + 1
            while i < len(body) and body[i] != '"':
                i += 1
            value = body[value_start + 1 : i]
            pos = i + 1
        else:
            i = value_start
            while i < len(body) and body[i] != ",":
                i += 1
            value = body[value_start:i]
            pos = i
        fields[name] = latex_to_text(value.strip())
    return fields


def cited_keys() -> list[str]:
    text = PLAN_AUX.read_text(encoding="utf-8", errors="ignore")
    keys: list[str] = []
    for key in re.findall(r"\\abx@aux@cite\{0\}\{([^{}]+)\}", text):
        if key not in keys:
            keys.append(key)
    return keys


def load_zotero_json() -> list[dict]:
    if not ZOTERO_JSON.exists():
        return []
    return json.loads(ZOTERO_JSON.read_text(encoding="utf-8"))


def zotero_sqlite_records() -> list[dict]:
    if not ZOTERO_SQLITE.exists():
        return []
    try:
        con = sqlite3.connect(
            f"file:{ZOTERO_SQLITE.as_posix()}?mode=ro&immutable=1", uri=True
        )
    except sqlite3.Error:
        return []

    rows = con.execute(
        """
        SELECT i.itemID, i.key,
               MAX(CASE WHEN f.fieldName='title' THEN v.value END) AS title,
               MAX(CASE WHEN f.fieldName='date' THEN v.value END) AS item_date,
               MAX(CASE WHEN f.fieldName='abstractNote' THEN v.value END) AS abstractNote,
               MAX(CASE WHEN f.fieldName='DOI' THEN v.value END) AS doi,
               MAX(CASE WHEN f.fieldName='url' THEN v.value END) AS url,
               MAX(CASE WHEN f.fieldName='publicationTitle' THEN v.value END) AS publicationTitle
        FROM items i
        LEFT JOIN itemData d ON d.itemID=i.itemID
        LEFT JOIN fieldsCombined f ON f.fieldID=d.fieldID
        LEFT JOIN itemDataValues v ON v.valueID=d.valueID
        WHERE i.itemID NOT IN (SELECT itemID FROM deletedItems)
        GROUP BY i.itemID, i.key
        """
    ).fetchall()

    records: list[dict] = []
    for item_id, key, title, date, abstract, doi, url, pub in rows:
        if not title:
            continue
        creators = con.execute(
            """
            SELECT c.firstName, c.lastName
            FROM itemCreators ic
            JOIN creators c ON c.creatorID=ic.creatorID
            WHERE ic.itemID=?
            ORDER BY ic.orderIndex
            """,
            (item_id,),
        ).fetchall()
        authors = [
            " ".join(part for part in creator if part).strip() for creator in creators
        ]
        attachments = []
        for att_key, att_path in con.execute(
            """
            SELECT i.key, a.path
            FROM itemAttachments a
            JOIN items i ON i.itemID=a.itemID
            WHERE a.parentItemID=? AND a.contentType='application/pdf'
            """,
            (item_id,),
        ).fetchall():
            if not att_path:
                continue
            if att_path.startswith("storage:"):
                candidate = ZOTERO_SQLITE.parent / "storage" / att_key / att_path[8:]
            else:
                candidate = Path(att_path)
            if candidate.exists():
                attachments.append(str(candidate))
        records.append(
            {
                "title": title,
                "date": date or "",
                "abstractNote": abstract or "",
                "DOI": doi or "",
                "url": url or "",
                "publicationTitle": pub or "",
                "authors": authors,
                "attachments": attachments,
            }
        )
    con.close()
    return records


def best_zotero_match(entry: dict, json_items: list[dict], sqlite_items: list[dict]) -> dict:
    doi = (entry.get("doi") or "").lower().strip()
    title_norm = normalize(entry.get("title", ""))

    for item in json_items:
        if doi and (item.get("DOI") or "").lower().strip() == doi:
            return item
    for item in sqlite_items:
        if doi and (item.get("DOI") or "").lower().strip() == doi:
            return item

    candidates = json_items + sqlite_items
    best_item: dict = {}
    best_score = 0.0
    title_tokens = set(title_norm.split())
    for item in candidates:
        other_tokens = set(normalize(item.get("title", "")).split())
        if not title_tokens or not other_tokens:
            continue
        score = len(title_tokens & other_tokens) / max(len(title_tokens), 1)
        if score > best_score:
            best_score = score
            best_item = item
    return best_item if best_score >= 0.55 else {}


def all_pdf_candidates(sqlite_items: list[dict]) -> list[Path]:
    paths: list[Path] = []
    for folder in PDF_SEARCH_DIRS:
        if folder.exists():
            paths.extend(folder.rglob("*.pdf"))
    for item in sqlite_items:
        for attachment in item.get("attachments", []):
            p = Path(attachment)
            if p.exists():
                paths.append(p)
    seen = set()
    unique: list[Path] = []
    for p in paths:
        key = str(p.resolve()).lower()
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def find_pdf(key: str, entry: dict, pdfs: list[Path], sqlite_match: dict) -> Path | None:
    hint = PDF_HINTS.get(key)
    if hint:
        hint_norm = normalize(hint)
        for pdf in pdfs:
            if hint_norm and hint_norm in normalize(pdf.name):
                return pdf
    for attachment in sqlite_match.get("attachments", []):
        p = Path(attachment)
        if p.exists():
            return p

    doi_norm = normalize(entry.get("doi", ""))
    if doi_norm:
        doi_tail = doi_norm.split()[-1] if doi_norm.split() else doi_norm
        for pdf in pdfs:
            name_norm = normalize(pdf.name)
            if doi_tail and doi_tail in name_norm:
                return pdf

    title_tokens = {
        t for t in normalize(entry.get("title", "")).split() if len(t) >= 5
    }
    if not title_tokens:
        return None
    best_pdf = None
    best_score = 0.0
    for pdf in pdfs:
        name_tokens = set(normalize(pdf.stem).split())
        score = len(title_tokens & name_tokens) / max(len(title_tokens), 1)
        if score > best_score:
            best_pdf = pdf
            best_score = score
    return best_pdf if best_score >= 0.35 else None


def category_for(key: str, entry_type: str) -> str:
    if key in STANDARD_KEYS:
        return "Norma, estandar, guia o informe tecnico"
    if entry_type in RESEARCH_TYPES:
        return "Paper academico usado"
    if entry_type in MANUAL_TYPES:
        return "Manual, libro o fuente tecnica"
    return "Fuente bibliografica usada"


def split_top_level_rows(table_body: str) -> list[str]:
    rows: list[str] = []
    current: list[str] = []
    i = 0
    while i < len(table_body):
        if table_body[i : i + 2] == r"\\":
            row = "".join(current).strip()
            if row:
                rows.append(row)
            current = []
            i += 2
        else:
            current.append(table_body[i])
            i += 1
    trailing = "".join(current).strip()
    if trailing:
        rows.append(trailing)
    return rows


def split_cells(row: str) -> list[str]:
    cells: list[str] = []
    current: list[str] = []
    brace_level = 0
    for ch in row:
        if ch == "{":
            brace_level += 1
        elif ch == "}":
            brace_level = max(0, brace_level - 1)
        if ch == "&" and brace_level == 0:
            cells.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    cells.append("".join(current).strip())
    return [latex_to_text(cell) for cell in cells]


def extract_matrix_table(label: str) -> tuple[list[str], list[list[str]]]:
    text = PLAN_TEX.read_text(encoding="utf-8")
    label_pos = text.index(r"\label{" + label + "}")
    begin_pos = text.rfind(r"\begin{longtable}", 0, label_pos)
    end_pos = text.index(r"\end{longtable}", label_pos)
    segment = text[begin_pos:end_pos]
    body = segment.split(r"\endlastfoot", 1)[1]
    for token in [r"\toprule", r"\midrule", r"\bottomrule"]:
        body = body.replace(token, "")
    raw_rows = split_top_level_rows(body)
    parsed: list[list[str]] = []
    for row in raw_rows:
        if any(
            token in row
            for token in [
                r"\caption",
                r"\label",
                r"\endfirsthead",
                r"\endhead",
                r"\endfoot",
                r"\multicolumn",
            ]
        ):
            continue
        cells = split_cells(row)
        if len(cells) > 1:
            parsed.append(cells)
    if label.endswith("-a"):
        headers = [
            "Fase DSR / actividad",
            "Problema",
            "Objetivo",
            "Producto verificable",
        ]
    else:
        headers = [
            "Fase DSR / actividad",
            "Problema",
            "Insumo o entrada del problema",
            "Configuracion del insumo",
            "Producto o resultado esperado",
            "Hipotesis",
            "Evidencia o indicadores de evaluacion",
            "Condiciones de control",
        ]
    return headers, parsed


def style_sheet(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    thin = Side(style="thin", color="B7BDC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="7C1F2B")
    body_fill = PatternFill("solid", fgColor="F7F9FB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = body_fill
            cell.font = Font(size=9)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 90


def write_matrix_excel(path: Path) -> None:
    headers_a, rows_a = extract_matrix_table("tab:matriz-consistencia-a")
    headers_b, rows_b = extract_matrix_table("tab:matriz-consistencia-b")

    wb = Workbook()
    wb.properties.title = "Matriz de consistencia extraida del documento final"
    wb.properties.creator = "Codex"

    ws = wb.active
    ws.title = "Parte A documento"
    ws.append(headers_a)
    for row in rows_a:
        ws.append(row)
    style_sheet(ws, [28, 55, 50, 44])

    ws = wb.create_sheet("Parte B documento")
    ws.append(headers_b)
    for row in rows_b:
        ws.append(row)
    style_sheet(ws, [28, 36, 28, 36, 34, 48, 38, 38])

    ws = wb.create_sheet("Matriz unificada")
    unified_headers = headers_a + [
        h
        for h in headers_b[2:]
        if h not in {"Fase DSR / actividad", "Problema"}
    ]
    ws.append(unified_headers)
    b_by_phase = {row[0]: row for row in rows_b}
    for row_a in rows_a:
        row_b = b_by_phase.get(row_a[0], [])
        ws.append(row_a + row_b[2:])
    style_sheet(ws, [28, 44, 38, 34, 24, 32, 30, 42, 34, 34])

    ws = wb.create_sheet("Nota")
    ws.append(["Campo", "Valor"])
    ws.append(["Fuente", str(PLAN_TEX.relative_to(ROOT))])
    ws.append(["Criterio", "Contenido extraido de las tablas finales del Anexo 5.1 del documento, sin usar la matriz previa de la carpeta matriz_consistencia."])
    ws.append(["Filas parte A", len(rows_a)])
    ws.append(["Filas parte B", len(rows_b)])
    style_sheet(ws, [24, 100])
    wb.save(path)


def write_references_excel(path: Path, records: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fuentes usadas"
    headers = [
        "Clave",
        "Categoria",
        "Tipo BibTeX",
        "Titulo",
        "Autores",
        "Ano",
        "Fuente",
        "DOI",
        "URL",
        "PDF encontrado",
        "Archivo copiado",
        "Resumen local",
        "Aporte a la tesis",
    ]
    ws.append(headers)
    for record in records:
        ws.append([record.get(h, "") for h in headers])
    style_sheet(ws, [18, 30, 16, 55, 36, 10, 36, 28, 42, 18, 46, 72, 72])
    wb.save(path)


def write_inventory_excel(path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(["Seccion", "Archivo o carpeta", "Descripcion"])
    for row in rows:
        ws.append(row)
    style_sheet(ws, [32, 60, 70])
    wb.save(path)


def copy_examples(out_dir: Path) -> list[list[str]]:
    target = out_dir / "02_ejemplos_datos_objeto_estudio"
    copied: list[list[str]] = []
    suffixes = {".csv", ".txt", ".md", ".png"}
    for root in [ROOT / "pinn_cables" / "data", ROOT / "examples"]:
        if not root.exists():
            continue
        for src in root.rglob("*"):
            if not src.is_file() or src.suffix.lower() not in suffixes:
                continue
            if "checkpoints" in src.parts:
                continue
            if src.suffix.lower() == ".png" and "results" not in normalize(str(src)):
                continue
            rel = src.relative_to(ROOT)
            dst = target / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            copied.append(["Ejemplos de datos", str(dst.relative_to(out_dir)), "CSV, reportes o figuras de ejemplo del objeto de estudio."])
    return copied


def main() -> None:
    out_dir = unique_output_dir()
    papers_dir = out_dir / "01_papers_usados"
    normas_dir = out_dir / "03_normas_estandares_guias_manuales"
    matrix_dir = out_dir / "05_matriz_consistencia"
    papers_dir.mkdir(parents=True, exist_ok=True)
    normas_dir.mkdir(parents=True, exist_ok=True)
    matrix_dir.mkdir(parents=True, exist_ok=True)

    bib_entries = parse_bib(PLAN_BIB)
    keys = cited_keys()
    json_items = load_zotero_json()
    sqlite_items = zotero_sqlite_records()
    pdfs = all_pdf_candidates(sqlite_items)

    records: list[dict] = []
    inventory: list[list[str]] = []
    used_names: set[str] = set()

    for key in keys:
        entry = bib_entries.get(key, {"ID": key, "ENTRYTYPE": "desconocido"})
        entry_type = entry.get("ENTRYTYPE", "")
        title = entry.get("title", key)
        category = category_for(key, entry_type)
        zotero_match = best_zotero_match(entry, json_items, sqlite_items)
        pdf = find_pdf(key, entry, pdfs, zotero_match)

        if category == "Paper academico usado":
            copy_dir = papers_dir
        else:
            copy_dir = normas_dir

        copied_name = ""
        if pdf:
            copied_name = safe_filename(title)
            if copied_name.lower() in used_names:
                copied_name = safe_filename(f"{title} - {key}")
            used_names.add(copied_name.lower())
            shutil.copy2(pdf, copy_dir / copied_name)
            inventory.append([category, str((copy_dir / copied_name).relative_to(out_dir)), f"Fuente original: {pdf}"])

        authors = entry.get("author", "")
        if not authors and zotero_match.get("authors"):
            authors = "; ".join(zotero_match.get("authors", []))
        abstract = zotero_match.get("abstractNote", "") or "Resumen/abstract local no disponible; revisar la fuente PDF o bibliografica."
        records.append(
            {
                "Clave": key,
                "Categoria": category,
                "Tipo BibTeX": entry_type,
                "Titulo": title,
                "Autores": authors,
                "Ano": entry.get("year", "") or entry.get("date", ""),
                "Fuente": entry.get("journaltitle", "")
                or entry.get("booktitle", "")
                or entry.get("institution", "")
                or entry.get("publisher", ""),
                "DOI": entry.get("doi", "") or zotero_match.get("DOI", ""),
                "URL": entry.get("url", "") or zotero_match.get("url", ""),
                "PDF encontrado": "Si" if pdf else "No",
                "Archivo copiado": str((copy_dir / copied_name).relative_to(out_dir)) if copied_name else "",
                "Resumen local": abstract,
                "Aporte a la tesis": APORTES.get(key, "Fuente usada como soporte bibliografico del documento final."),
            }
        )

    references_xlsx = papers_dir / "papers_usados_resumenes.xlsx"
    write_references_excel(references_xlsx, records)
    inventory.append(["Excel de papers y fuentes", str(references_xlsx.relative_to(out_dir)), "Metadatos, resumen local y aporte a la tesis de las fuentes citadas en el documento final."])

    matrix_xlsx = matrix_dir / "Matriz_de_consistencia_documento_final.xlsx"
    write_matrix_excel(matrix_xlsx)
    inventory.append(["Matriz de consistencia", str(matrix_xlsx.relative_to(out_dir)), "Excel generado desde las tablas finales del anexo del documento LaTeX."])

    inventory.extend(copy_examples(out_dir))

    readme = out_dir / "LEEME.txt"
    readme.write_text(
        "\n".join(
            [
                "Paquete listo para subir manualmente a Google Drive.",
                "",
                "Contenido:",
                "1. 01_papers_usados: PDFs academicos citados y Excel con datos basicos, resumen local y aporte a la tesis.",
                "2. 02_ejemplos_datos_objeto_estudio: CSV, reportes y figuras de ejemplos del objeto de estudio tomados de examples/ y pinn_cables/data.",
                "3. 03_normas_estandares_guias_manuales: normas, estandares, guias, manuales e informes tecnicos citados con PDF local encontrado.",
                "5. 05_matriz_consistencia: Excel de la matriz extraida del documento final Plan/plan_tesis_cables_pinn.tex.",
                "",
                "Nota: si una fuente citada no tenia PDF local detectable, aparece marcada como 'PDF encontrado = No' en 01_papers_usados/papers_usados_resumenes.xlsx.",
            ]
        ),
        encoding="utf-8",
    )
    inventory.append(["Indice", str(readme.relative_to(out_dir)), "Descripcion breve del paquete."])

    inventory_xlsx = out_dir / "00_indice_entrega.xlsx"
    write_inventory_excel(inventory_xlsx, inventory)

    print(out_dir)
    print(f"Fuentes citadas: {len(keys)}")
    print(f"PDFs copiados: {sum(1 for r in records if r['PDF encontrado'] == 'Si')}")
    print(f"Ejemplos copiados: {sum(1 for r in inventory if r[0] == 'Ejemplos de datos')}")


if __name__ == "__main__":
    main()
