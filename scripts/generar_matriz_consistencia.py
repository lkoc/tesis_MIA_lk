from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
MATRIX_DIR = ROOT / "matriz_consistencia"
OUTPUT = MATRIX_DIR / "Matriz_de_Consistencia_Tesis_Cables_Heterogeneidad_Termica.xlsx"
ZOTERO_DB = Path.home() / "Zotero" / "zotero.sqlite"

TITLE = (
    "Efecto de la heterogeneidad térmica del entorno sobre la temperatura del "
    "conductor en cables eléctricos subterráneos en centrales solares y eólicas: "
    "modelado numérico 2D"
)
AUTHORS = "Luis Koc y Herbert Meléndez"

BURGUNDY = "8F1D2C"
BURGUNDY_DARK = "65131F"
TEAL = "167D7F"
GREEN = "4F7942"
MUSTARD = "D5A021"
PALE_RED = "F7E8EA"
PALE_TEAL = "E5F3F3"
PALE_GREEN = "EAF1E7"
PALE_GOLD = "FAF2D9"
PALE_BLUE = "E9F0F7"
LIGHT_GRAY = "E7E9EC"
MID_GRAY = "A7ADB4"
DARK = "1F2933"
WHITE = "FFFFFF"

THIN_GRAY = Side(style="thin", color="A7ADB4")
MEDIUM_BURGUNDY = Side(style="medium", color=BURGUNDY)
ALL_THIN = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def set_cell(
    cell,
    *,
    value=None,
    font=None,
    fill=None,
    alignment=None,
    border=None,
):
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def style_range(ws, cell_range, *, font=None, fill=None, alignment=None, border=None):
    for row in ws[cell_range]:
        for cell in row:
            set_cell(
                cell,
                font=font,
                fill=fill,
                alignment=alignment,
                border=border,
            )


def configure_page(ws, *, landscape=True, paper="8", fit_height=1):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = paper
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.oddFooter.center.text = "Matriz de consistencia - " + AUTHORS
    ws.oddFooter.right.text = "Página &P de &N"


def add_title(ws, last_col, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    set_cell(
        ws.cell(1, 1),
        value="MATRIZ DE CONSISTENCIA",
        font=Font(name="Arial", size=15, bold=True, color=WHITE),
        fill=PatternFill("solid", fgColor=BURGUNDY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=Border(bottom=MEDIUM_BURGUNDY),
    )
    set_cell(
        ws.cell(2, 1),
        value=TITLE,
        font=Font(name="Arial", size=11, bold=True, color=DARK),
        fill=PatternFill("solid", fgColor=PALE_RED),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=ALL_THIN,
    )
    set_cell(
        ws.cell(3, 1),
        value=subtitle,
        font=Font(name="Arial", size=9, italic=True, color=DARK),
        fill=PatternFill("solid", fgColor=LIGHT_GRAY),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=ALL_THIN,
    )
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 30


def build_official_matrix(wb):
    ws = wb.active
    ws.title = "Matriz UNI"
    add_title(
        ws,
        6,
        "Maestría en Inteligencia Artificial - FIIS/UNI | Estructura conforme a "
        "las Guías Metodológicas UNI N.° 01 y N.° 02 | Autores: "
        f"{AUTHORS} | Elaboración: {date.today().strftime('%d/%m/%Y')}",
    )

    headers = [
        "Título / problema de estudio",
        "Problemas",
        "Objetivos",
        "Hipótesis",
        "Variables y dimensiones",
        "Diseño metodológico",
    ]
    for col, text in enumerate(headers, 1):
        set_cell(
            ws.cell(4, col),
            value=text,
            font=Font(name="Arial", size=9, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=BURGUNDY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=ALL_THIN,
        )
    ws.row_dimensions[4].height = 36

    problem_study = (
        f"TÍTULO\n{TITLE}\n\n"
        "PROBLEMA DE ESTUDIO\n"
        "Los métodos normativos y simplificados suelen representar el suelo y el "
        "relleno con propiedades térmicas homogéneas. En instalaciones reales, la "
        "conductividad térmica varía espacialmente por estratos, humedad, secado, "
        "bedding y backfill; esta heterogeneidad puede alterar los puntos calientes, "
        "la temperatura máxima del conductor y la ampacidad estimada."
    )
    general_problem = (
        "PROBLEMA GENERAL\n"
        "¿Cuál es el efecto de la heterogeneidad térmica espacial del entorno, "
        "representada mediante k(x,y), sobre la temperatura máxima del conductor y "
        "la ampacidad de cables eléctricos subterráneos en escenarios representativos "
        "de centrales solares y eólicas, según un modelo numérico 2D previamente "
        "verificado y validado?"
    )
    general_objective = (
        "OBJETIVO GENERAL\n"
        "Determinar y cuantificar el efecto de la heterogeneidad térmica espacial del "
        "suelo, bedding y backfill sobre la temperatura máxima del conductor y la "
        "ampacidad de cables eléctricos subterráneos en centrales solares y eólicas, "
        "mediante la aplicación de un modelo numérico 2D previamente verificado y "
        "validado."
    )
    general_hypothesis = (
        "HIPÓTESIS GENERAL\n"
        "La heterogeneidad térmica espacial modifica de forma sistemática la respuesta "
        "térmica respecto de un entorno homogéneo equivalente: las zonas de baja k "
        "próximas al cable elevan Tmax y reducen la ampacidad, mientras que un backfill "
        "de alta k reduce Tmax; la magnitud del efecto aumenta con el contraste de k, "
        "la extensión de la zona heterogénea y su proximidad al conductor."
    )
    general_variables = (
        "VARIABLE INDEPENDIENTE (X)\n"
        "Heterogeneidad térmica espacial del entorno, k(x,y).\n"
        "Dimensiones: magnitud de k; contraste Ck=kmax/kmin; dispersión CVk; patrón "
        "espacial; proximidad dh; extensión fh; tipo de material/estado hídrico.\n\n"
        "VARIABLE DEPENDIENTE PRINCIPAL (Y1)\n"
        "Temperatura del conductor.\n"
        "Dimensiones: Tmax, ΔTmax, campo T(x,y), gradiente y localización del hotspot.\n\n"
        "VARIABLE DEPENDIENTE DERIVADA (Y2)\n"
        "Ampacidad.\n"
        "Dimensiones: Imax para Tlim, derating y margen térmico."
    )
    methodology = (
        "Enfoque: cuantitativo.\n"
        "Tipo: aplicada/tecnológica.\n"
        "Nivel: explicativo-predictivo.\n"
        "Diseño: experimental numérico in silico, controlado, factorial y de "
        "sensibilidad; régimen 2D estacionario como núcleo y casos transitorios "
        "seleccionados si el alcance lo permite.\n"
        "Métodos: hipotético-deductivo, modelado numérico y Design Science Research "
        "para construir/evaluar el artefacto computacional.\n"
        "Unidad de análisis: sección transversal 2D de una instalación de cables "
        "subterráneos.\n"
        "Población: posibles secciones 2D de circuitos subterráneos en centrales "
        "renovables. Muestra: no probabilística e intencional, formada por benchmarks "
        "y escenarios factoriales basados en IEC 60287/60853, Aras (2005), Kim et al. "
        "(2025) y configuraciones representativas de circuitos renovables.\n"
        "Técnicas/instrumentos: revisión documental; matriz paramétrica; solver PINN "
        "2D en Python; referencia FEM/analítica/normativa; archivos CSV y registros "
        "reproducibles.\n"
        "Control de calidad: verificación numérica frente a soluciones analíticas o "
        "FEM con convergencia demostrada; validación física frente a datos "
        "experimentales o publicados cuando estén disponibles. IEC se usa como línea "
        "base normativa, no como verdad de referencia para campos heterogéneos.\n"
        "Análisis: error de Tmax y T(x,y), residuos PDE/BC, balance de energía, "
        "incertidumbre del modelo, ΔI%, curvas de respuesta, regresión/sensibilidad y "
        "repetición con semillas. La incertidumbre numérica debe ser menor que el "
        "efecto térmico que se pretende atribuir a k(x,y)."
    )

    specifics = [
        (
            "PE1. ¿Cómo influyen la magnitud y el contraste de la conductividad "
            "térmica k(x,y) en la temperatura máxima del conductor y el campo "
            "térmico cuando se mantienen constantes la geometría, la carga y la "
            "disposición espacial?",
            "OE1. Determinar el efecto de la magnitud y el contraste de k(x,y) sobre "
            "Tmax y T(x,y) mediante comparaciones controladas con un entorno "
            "homogéneo equivalente.",
            "HE1. La disminución de k en el entorno y el aumento del contraste térmico "
            "incrementarán Tmax y los gradientes térmicos respecto del caso homogéneo "
            "equivalente.",
            "X: kmin, kmedia, Ck y CVk.\n"
            "Y1: Tmax, ΔTmax, T(x,y) y gradiente térmico.\n"
            "Controles: geometría, carga, patrón espacial, profundidad y BC.",
        ),
        (
            "PE2. ¿Cómo varía Tmax con el contraste, patrón espacial, extensión y "
            "proximidad al conductor de las regiones con distinta conductividad "
            "térmica?",
            "OE2. Cuantificar la variación de Tmax y del campo T(x,y) frente a "
            "escenarios controlados de heterogeneidad k(x,y), aislando el efecto de "
            "sus dimensiones espaciales.",
            "HE2. A igualdad de geometría, carga y condiciones de contorno, una "
            "región de baja k más extensa y próxima al cable incrementará Tmax; el "
            "incremento crecerá con Ck y CVk respecto del caso homogéneo equivalente.",
            "X: Ck, CVk, patrón, dh y fh.\n"
            "Y1: Tmax, ΔTmax, T(x,y), gradiente térmico y hotspot.\n"
            "Controles: geometría, pérdidas, corriente, profundidad, dominio y BC.",
        ),
        (
            "PE3. ¿En qué medida la heterogeneidad térmica modifica la corriente "
            "máxima admisible al imponer el límite térmico del conductor?",
            "OE3. Determinar el cambio de ampacidad y el margen térmico para cada "
            "escenario de heterogeneidad, incluyendo suelo seco y backfill mejorado.",
            "HE3. Las heterogeneidades de baja k reducirán Imax y el margen térmico; "
            "un bedding/backfill de alta k incrementará Imax respecto de la "
            "representación homogénea de referencia.",
            "X: escenario k(x,y) y propiedades del entorno.\n"
            "Y2: Imax, ΔI%, derating y margen a Tlim.\n"
            "Y1 auxiliar: Tmax para la iteración de corriente.",
        ),
        (
            "PE4. ¿En qué condiciones los métodos homogéneos simplificados presentan "
            "la mayor discrepancia frente al modelo 2D heterogéneo?",
            "OE4. Comparar las estimaciones térmicas y de ampacidad del modelo 2D con "
            "IEC 60287/60853 y referencias equivalentes, e identificar las "
            "configuraciones críticas para centrales solares y eólicas.",
            "HE4. La discrepancia de Tmax y ampacidad será mayor cuando exista alto "
            "contraste de k y una zona de baja conductividad cercana al cable; el uso "
            "de una k promedio ocultará hotspots y tenderá a sobreestimar la ampacidad.",
            "Factor comparativo: método homogéneo vs. modelo 2D heterogéneo.\n"
            "X moderadora: intensidad/localización de k(x,y).\n"
            "Respuestas: brecha térmica, error de ampacidad y clasificación de riesgo.",
        ),
    ]

    ws.merge_cells("A5:A9")
    ws.merge_cells("F5:F9")
    ws["A5"] = problem_study
    ws["B5"] = general_problem
    ws["C5"] = general_objective
    ws["D5"] = general_hypothesis
    ws["E5"] = general_variables
    ws["F5"] = methodology

    for idx, row in enumerate(specifics, start=6):
        ws.cell(idx, 2).value = row[0]
        ws.cell(idx, 3).value = row[1]
        ws.cell(idx, 4).value = row[2]
        ws.cell(idx, 5).value = row[3]

    general_fill = PatternFill("solid", fgColor=PALE_GOLD)
    specific_fills = [
        PatternFill("solid", fgColor=PALE_BLUE),
        PatternFill("solid", fgColor=PALE_TEAL),
        PatternFill("solid", fgColor=PALE_GREEN),
        PatternFill("solid", fgColor=PALE_RED),
    ]
    style_range(
        ws,
        "A5:F9",
        font=Font(name="Arial", size=8.2, color=DARK),
        alignment=Alignment(vertical="top", wrap_text=True),
        border=ALL_THIN,
    )
    style_range(ws, "B5:E5", fill=general_fill)
    ws["A5"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws["F5"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws["A5"].font = Font(name="Arial", size=8.5, bold=True, color=DARK)
    ws["F5"].font = Font(name="Arial", size=8.2, color=DARK)
    for idx, fill in enumerate(specific_fills, start=6):
        style_range(ws, f"B{idx}:E{idx}", fill=fill)

    ws.row_dimensions[5].height = 145
    for row in range(6, 10):
        ws.row_dimensions[row].height = 108
    widths = {"A": 26, "B": 31, "C": 31, "D": 34, "E": 29, "F": 37}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = "B4:E9"
    ws.print_title_rows = "1:4"
    ws.print_area = "A1:F9"
    configure_page(ws, landscape=True, paper="8", fit_height=1)
    return ws


def build_activity_matrix(wb):
    ws = wb.create_sheet("Matriz por actividades")
    add_title(
        ws,
        10,
        "Adaptación del Excel de ejemplo: fases, productos verificables, variables, "
        "hipótesis, métricas y controles",
    )
    headers = [
        "Fase / actividad",
        "Problema",
        "Objetivo",
        "Producto verificable",
        "Variable independiente",
        "Estados de la variable independiente",
        "Variable dependiente",
        "Hipótesis / resultado esperado",
        "Métricas",
        "Variables de control",
    ]
    for col, text in enumerate(headers, 1):
        set_cell(
            ws.cell(4, col),
            value=text,
            font=Font(name="Arial", size=8.5, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=TEAL),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=ALL_THIN,
        )

    rows = [
        [
            "1. Revisión y especificación",
            "Los supuestos y limitaciones de IEC, Neher-McGrath y modelos 2D están "
            "dispersos; la biblioteca debe conectarse con el diseño experimental.",
            "Sistematizar antecedentes, categorías Zotero, parámetros, vacíos y "
            "criterios de comparación.",
            "Matriz documental; taxonomía de escenarios; ficha de parámetros y "
            "trazabilidad de fuentes.",
            "No aplica como variable causal; fase conceptual.",
            "7 temas Zotero: PINN; transferencia de calor; suelos; FEM; cables; DTR; "
            "normas. Tipos: estado del arte, aporte, aplicación, normativa y guía.",
            "Cobertura y trazabilidad de la especificación.",
            "La revisión estructurada permitirá traducir los vacíos reportados en "
            "escenarios medibles y criterios de validación.",
            "% fuentes clasificadas; cobertura temática; parámetros con fuente; "
            "duplicados/conflictos resueltos.",
            "Fecha de corte; criterios de inclusión; versión de normas; idioma; "
            "alcance 2D.",
        ],
        [
            "2. Construcción del modelo 2D",
            "Se requiere un artefacto reproducible que represente dominios "
            "multimaterial y k(x,y) sin perder consistencia física.",
            "Implementar la ecuación de calor 2D, pérdidas internas y condiciones de "
            "contorno para cables enterrados.",
            "Solver PINN 2D; configuraciones CSV; geometrías; pruebas unitarias; "
            "salidas de campo térmico.",
            "Representación del dominio y configuración numérica.",
            "Homogéneo; multicapa; inclusión localizada; bedding/backfill; régimen "
            "estacionario y casos transitorios seleccionados.",
            "Consistencia física y estabilidad numérica.",
            "El solver representará las interfaces y condiciones de contorno con "
            "residuos controlados y resultados reproducibles.",
            "Residual PDE/BC; balance de energía; convergencia; dispersión entre "
            "semillas; tiempo de cómputo.",
            "Arquitectura PINN; puntos de colocación; seed; precisión; dominio; "
            "tolerancias; hardware.",
        ],
        [
            "3. Verificación numérica y validación física",
            "Necesidad metodológica: antes de atribuir cambios de Tmax a k(x,y), debe "
            "demostrarse que la discrepancia e incertidumbre del modelo son menores "
            "que el efecto térmico analizado.",
            "Verificar que el solver resuelva correctamente la ecuación de calor y "
            "validar que represente la respuesta física dentro de un dominio de "
            "aplicación explícito.",
            "Informe V&V: pruebas analíticas/FEM, contraste con datos experimentales o "
            "publicados, presupuesto de error y dominio de validez.",
            "No es una variable causal de la investigación; es un control "
            "metodológico.",
            "Verificación: solución analítica o manufacturada y FEM con convergencia. "
            "Validación: datos experimentales/publicados. IEC: comparación normativa.",
            "Discrepancia e incertidumbre del modelo.",
            "No corresponde una hipótesis científica adicional; se aplica un criterio "
            "de aceptación previo a los experimentos de heterogeneidad.",
            "Error de Tmax y T(x,y); residuos PDE/BC; balance de energía; convergencia; "
            "sesgo; dispersión entre semillas; Umodelo/|ΔTefecto|.",
            "Mismos materiales, pérdidas, geometría, dominio y BC; referencia "
            "independiente; malla FEM convergente; semillas registradas.",
        ],
        [
            "4. Experimento de heterogeneidad",
            "No está cuantificado cuánto cambian Tmax y el hotspot cuando k varía en "
            "magnitud y espacio.",
            "Estimar los efectos principales e interacciones de Ck, CVk, patrón, dh y "
            "fh sobre la respuesta térmica.",
            "Base de escenarios; mapas T(x,y); curvas de respuesta; análisis de "
            "sensibilidad.",
            "Heterogeneidad térmica k(x,y).",
            "Homogéneo; estratificado; zona seca/baja k; bedding diferenciado; "
            "backfill/PAC de alta k; contrastes bajo, medio y alto.",
            "Tmax y campo térmico.",
            "Las regiones de baja k próximas y extensas elevarán Tmax; el efecto "
            "crecerá con Ck y CVk.",
            "Tmax; ΔTmax; hotspot; gradiente; MAE espacial vs. base; sensibilidad "
            "normalizada.",
            "Corriente; pérdidas; geometría; profundidad; separación; Tamb; BC; "
            "propiedades del cable.",
        ],
        [
            "5. Ampacidad y brecha normativa",
            "La k promedio puede ocultar hotspots y sesgar la corriente admisible en "
            "condiciones heterogéneas.",
            "Calcular Imax y comparar el modelo 2D con IEC 60287/60853 en escenarios "
            "representativos de plantas solares y eólicas.",
            "Curvas I-Tmax; tabla de derating; mapa de riesgo y condiciones críticas.",
            "Escenario k(x,y) y método térmico.",
            "Modelo 2D heterogéneo vs. método homogéneo; suelo seco vs. húmedo; "
            "backfill convencional vs. mejorado.",
            "Ampacidad y brecha de estimación.",
            "La baja k cercana reducirá Imax y ampliará la discrepancia del método "
            "homogéneo; la alta k del backfill tendrá el efecto opuesto.",
            "Imax; ΔI%; margen a Tlim; ΔT entre métodos; clasificación de riesgo.",
            "Tlim; perfil de carga; criterio de pérdidas; geometría; BC; método de "
            "iteración de corriente.",
        ],
        [
            "6. Síntesis y reproducibilidad",
            "Los resultados deben ser auditables, repetibles y traducibles en "
            "recomendaciones de ingeniería.",
            "Consolidar evidencia, limitaciones, recomendaciones y paquete "
            "reproducible de la investigación.",
            "Capítulos de resultados; repositorio; datos y configuraciones; matriz "
            "final; guía de reproducción.",
            "No aplica como variable causal; fase de cierre.",
            "Resultados confirmatorios; casos límite; fallos de convergencia; "
            "análisis de robustez.",
            "Calidad de evidencia y reproducibilidad.",
            "La trazabilidad completa permitirá separar hallazgos físicos, error "
            "numérico y variabilidad de entrenamiento.",
            "% casos reproducibles; scripts ejecutables; integridad de datos; "
            "cobertura de limitaciones.",
            "Versiones de código y dependencias; seeds; hardware; criterios de "
            "exclusión; fecha de corte.",
        ],
    ]
    for r_idx, values in enumerate(rows, start=5):
        for c_idx, value in enumerate(values, start=1):
            set_cell(
                ws.cell(r_idx, c_idx),
                value=value,
                font=Font(name="Arial", size=8, color=DARK),
                fill=PatternFill(
                    "solid", fgColor=WHITE if r_idx % 2 else "F5F7F8"
                ),
                alignment=Alignment(vertical="top", wrap_text=True),
                border=ALL_THIN,
            )
        ws.cell(r_idx, 1).font = Font(name="Arial", size=8, bold=True, color=BURGUNDY)
        ws.row_dimensions[r_idx].height = 118
    widths = [17, 25, 25, 23, 20, 25, 18, 27, 22, 24]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.row_dimensions[4].height = 45
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = "A4:J10"
    ws.print_title_rows = "1:4"
    ws.print_area = "A1:J10"
    configure_page(ws, landscape=True, paper="8", fit_height=2)
    return ws


def build_operationalization(wb):
    ws = wb.create_sheet("Operacionalización")
    add_title(
        ws,
        9,
        "Definiciones, dimensiones, indicadores, unidades, instrumentos y criterios "
        "de análisis para las variables del estudio | Hoja complementaria; las guías "
        "UNI la exigen formalmente solo para doctorado",
    )
    headers = [
        "Variable",
        "Tipo",
        "Definición conceptual",
        "Definición operacional",
        "Dimensión",
        "Indicador",
        "Unidad / escala",
        "Instrumento / fuente",
        "Criterio de análisis",
    ]
    for col, text in enumerate(headers, 1):
        set_cell(
            ws.cell(4, col),
            value=text,
            font=Font(name="Arial", size=8.5, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=GREEN),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=ALL_THIN,
        )

    rows = [
        [
            "Heterogeneidad térmica del entorno, k(x,y)",
            "Independiente",
            "Variación espacial de las propiedades que gobiernan la conducción de "
            "calor en suelo, bedding y backfill.",
            "Campo 2D de conductividad asignado por regiones o función espacial, "
            "manteniendo constantes los demás factores del escenario.",
            "Magnitud",
            "kmin; kmax; kmedia",
            "W/(m·K), razón",
            "Matriz paramétrica; literatura; IEC; configuración CSV",
            "Comparar niveles y verificar plausibilidad física.",
        ],
        [
            "Heterogeneidad térmica del entorno, k(x,y)",
            "Independiente",
            "Variación espacial de la capacidad conductora del entorno.",
            "Contraste y dispersión del campo k(x,y).",
            "Intensidad",
            "Ck=kmax/kmin; CVk=σk/kmedia",
            "Adimensional",
            "Script de generación y auditoría de escenarios",
            "Efectos principales; regresión; sensibilidad normalizada.",
        ],
        [
            "Heterogeneidad térmica del entorno, k(x,y)",
            "Independiente",
            "Configuración espacial de materiales y estados del suelo.",
            "Clasificación de cada mapa de material.",
            "Patrón espacial",
            "Homogéneo; estratificado; inclusión baja k; bedding; backfill/PAC",
            "Nominal",
            "Geometría 2D y mapa de materiales",
            "Contrastes respecto del caso homogéneo equivalente.",
        ],
        [
            "Heterogeneidad térmica del entorno, k(x,y)",
            "Independiente",
            "Posición relativa de la heterogeneidad respecto de la fuente térmica.",
            "Distancia mínima entre la región heterogénea y la superficie del cable.",
            "Proximidad",
            "dh",
            "m o múltiplos del diámetro",
            "Geometría computacional",
            "Curva efecto-distancia e interacción con Ck.",
        ],
        [
            "Heterogeneidad térmica del entorno, k(x,y)",
            "Independiente",
            "Tamaño relativo de la región con propiedad diferenciada.",
            "Área heterogénea dividida por el área de entorno evaluada.",
            "Extensión",
            "fh=Ah/Aentorno",
            "% o proporción",
            "Postproceso geométrico",
            "Curva dosis-respuesta e interacción con proximidad.",
        ],
        [
            "Temperatura del conductor",
            "Dependiente principal",
            "Respuesta térmica del conductor producida por pérdidas internas y "
            "transferencia de calor al entorno.",
            "Máximo de T en el dominio del conductor para cada simulación.",
            "Nivel térmico",
            "Tmax; Tmedia del conductor",
            "°C",
            "Solver 2D y postproceso",
            "Comparación con Tlim y con el escenario homogéneo.",
        ],
        [
            "Temperatura del conductor",
            "Dependiente principal",
            "Cambio atribuible a la representación heterogénea.",
            "Diferencia de Tmax frente a la línea base homogénea equivalente.",
            "Efecto térmico",
            "ΔTmax=Tmax,het-Tmax,hom",
            "K",
            "Tabla de resultados",
            "Magnitud, signo, monotonicidad e interacción de factores.",
        ],
        [
            "Temperatura del conductor",
            "Dependiente principal",
            "Distribución espacial de la temperatura y ubicación de la zona crítica.",
            "Campo discreto T(x,y), gradiente y coordenadas del máximo.",
            "Campo térmico",
            "MAE/RMSE; |∇T|max; xhot, yhot",
            "K; K/m; m",
            "Malla de evaluación y mapas 2D",
            "Comparación espacial, detección de hotspot y balance físico.",
        ],
        [
            "Ampacidad",
            "Dependiente derivada",
            "Corriente máxima que mantiene la temperatura del conductor dentro del "
            "límite admisible.",
            "Corriente obtenida iterativamente hasta cumplir Tmax=Tlim.",
            "Capacidad de corriente",
            "Imax",
            "A",
            "Rutina de búsqueda de corriente + solver 2D",
            "Comparación entre escenarios y con IEC 60287/60853.",
        ],
        [
            "Ampacidad",
            "Dependiente derivada",
            "Variación relativa de capacidad por efecto de la heterogeneidad.",
            "Cambio porcentual respecto del caso homogéneo o normativo.",
            "Derating / uprating",
            "ΔI%=100(Ihet-Iref)/Iref",
            "%",
            "Postproceso",
            "Signo y magnitud; identificación de condiciones críticas.",
        ],
        [
            "Exactitud del modelo",
            "Control metodológico",
            "Grado de concordancia entre el modelo propuesto y una referencia "
            "aceptada.",
            "Verificación del solver frente a soluciones conocidas y validación "
            "frente a evidencia física independiente.",
            "Verificación y validación",
            "Error de Tmax/campo; residuos; balance; sesgo; Umodelo/|ΔTefecto|",
            "K; %; adimensional",
            "Solución analítica/manufacturada; FEM convergente; datos "
            "experimentales/publicados",
            "Aceptar solo si hay convergencia y consistencia física, y si la "
            "incertidumbre del modelo es menor que el efecto de heterogeneidad "
            "evaluado; documentar el dominio de validez.",
        ],
        [
            "Carga y pérdidas",
            "Control",
            "Excitación térmica impuesta al cable.",
            "Corriente, perfil de carga y pérdidas fijados dentro de cada contraste.",
            "Operación",
            "I(t); Q; factor de carga",
            "A; W/m³; p.u.",
            "Entrada del modelo",
            "Mantener constante salvo experimento explícito.",
        ],
        [
            "Geometría e instalación",
            "Control / factor secundario",
            "Configuración física del cable y su instalación.",
            "Diámetros, capas, profundidad, separación y disposición parametrizados.",
            "Geometría",
            "D; z; s; configuración",
            "m; nominal",
            "Archivo de geometría",
            "Bloquear por caso o estudiar interacción por separado.",
        ],
        [
            "Condiciones de contorno y materiales",
            "Control",
            "Condiciones externas y propiedades no manipuladas.",
            "Tamb, dominio, BC, rho, cp y materiales del cable fijados por escenario.",
            "Entorno/modelo",
            "Tamb; tamaño de dominio; rho; cp; tipo BC",
            "°C; m; kg/m³; J/(kg·K)",
            "Configuración y fuentes documentales",
            "Igualdad estricta en comparaciones pareadas.",
        ],
        [
            "Configuración numérica",
            "Control",
            "Parámetros que afectan el error y la repetibilidad del solver.",
            "Arquitectura, pesos, colocación, seed, tolerancias y presupuesto de "
            "entrenamiento registrados.",
            "Reproducibilidad",
            "seed; épocas; puntos; tolerancia; runtime",
            "Conteo; s",
            "Logs y metadatos de ejecución",
            "Repetir semillas; reportar media, dispersión y fallos.",
        ],
    ]
    for r_idx, values in enumerate(rows, start=5):
        for c_idx, value in enumerate(values, 1):
            variable_type = values[1]
            if "Independiente" in variable_type:
                fill_color = PALE_TEAL
            elif "Dependiente" in variable_type:
                fill_color = PALE_GREEN
            elif "metodológica" in variable_type:
                fill_color = PALE_GOLD
            else:
                fill_color = "F3F4F6"
            set_cell(
                ws.cell(r_idx, c_idx),
                value=value,
                font=Font(name="Arial", size=8, color=DARK),
                fill=PatternFill("solid", fgColor=fill_color),
                alignment=Alignment(vertical="top", wrap_text=True),
                border=ALL_THIN,
            )
        ws.row_dimensions[r_idx].height = 64
    widths = [24, 18, 27, 28, 18, 23, 16, 24, 27]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.row_dimensions[4].height = 38
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4 + len(rows)}"
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:I{4 + len(rows)}"
    configure_page(ws, landscape=True, paper="8", fit_height=2)
    return ws


def zotero_connection():
    if not ZOTERO_DB.exists():
        return None
    return sqlite3.connect(f"file:{ZOTERO_DB.as_posix()}?mode=ro", uri=True)


def get_zotero_counts(con):
    total = con.execute(
        """
        SELECT COUNT(*)
        FROM items i
        JOIN itemTypes it ON it.itemTypeID=i.itemTypeID
        WHERE i.itemID NOT IN (SELECT itemID FROM deletedItems)
          AND it.typeName NOT IN ('attachment', 'note', 'annotation')
        """
    ).fetchone()[0]
    wanted = [
        ("tema-pinn", "PINN"),
        ("tema-transferencia-calor", "Transferencia de calor"),
        ("tema-suelo-backfill-humedad", "Suelo/backfill/humedad"),
        ("tema-fem-numerico", "FEM y métodos numéricos"),
        ("tema-cables-ampacidad", "Cables y ampacidad"),
        ("tema-dtr-cargas-ciclicas", "DTR y cargas cíclicas"),
        ("tema-normas-estandares", "Normas y estándares"),
        ("tipo-estado-del-arte", "Estado del arte"),
        ("tipo-aporte", "Aporte"),
        ("tipo-aplicacion", "Aplicación"),
        ("tipo-normativa", "Normativa"),
        ("tipo-guia-tecnica", "Guía técnica"),
        ("tipo-apoyo", "Apoyo"),
        ("tipo-fuente-clasica", "Fuente clásica"),
    ]
    counts = []
    for tag, label in wanted:
        count = con.execute(
            """
            SELECT COUNT(DISTINCT it.itemID)
            FROM itemTags it
            JOIN tags t ON t.tagID=it.tagID
            JOIN items i ON i.itemID=it.itemID
            WHERE i.itemID NOT IN (SELECT itemID FROM deletedItems)
              AND t.name=?
            """,
            (tag,),
        ).fetchone()[0]
        counts.append((label, tag, count))
    return total, counts


def get_zotero_items(con, item_ids):
    rows = []
    for item_id in item_ids:
        row = con.execute(
            """
            SELECT i.itemID, i.key,
                   MAX(CASE WHEN f.fieldName='title' THEN v.value END) AS title,
                   MAX(CASE WHEN f.fieldName='date' THEN v.value END) AS item_date,
                   GROUP_CONCAT(DISTINCT t.name) AS tags
            FROM items i
            LEFT JOIN itemData d ON d.itemID=i.itemID
            LEFT JOIN fields f ON f.fieldID=d.fieldID
            LEFT JOIN itemDataValues v ON v.valueID=d.valueID
            LEFT JOIN itemTags it ON it.itemID=i.itemID
            LEFT JOIN tags t ON t.tagID=it.tagID
            WHERE i.itemID=?
            GROUP BY i.itemID, i.key
            """,
            (item_id,),
        ).fetchone()
        if not row:
            continue
        creators = con.execute(
            """
            SELECT c.lastName, c.firstName
            FROM itemCreators ic
            JOIN creators c ON c.creatorID=ic.creatorID
            WHERE ic.itemID=?
            ORDER BY ic.orderIndex
            """,
            (item_id,),
        ).fetchall()
        author_text = "; ".join(
            ", ".join(part for part in creator if part) for creator in creators[:5]
        )
        rows.append(
            {
                "item_id": row[0],
                "key": row[1],
                "title": row[2] or "",
                "date": row[3] or "",
                "tags": row[4] or "",
                "authors": author_text,
            }
        )
    return rows


def build_sources(wb):
    ws = wb.create_sheet("Sustento Zotero")
    add_title(
        ws,
        7,
        "Auditoría de la base Zotero actual y selección de antecedentes que sustentan "
        "problema, variables, validación y metodología",
    )
    ws["A4"] = "Base Zotero"
    ws["B4"] = str(ZOTERO_DB)
    ws["A5"] = "Criterio"
    ws["B5"] = (
        "Se priorizan fuentes etiquetadas por tema y tipo en Zotero; la selección "
        "siguiente no sustituye la bibliografía completa."
    )
    style_range(
        ws,
        "A4:G5",
        font=Font(name="Arial", size=8.5, color=DARK),
        fill=PatternFill("solid", fgColor=LIGHT_GRAY),
        alignment=Alignment(vertical="center", wrap_text=True),
        border=ALL_THIN,
    )
    ws["A4"].font = Font(name="Arial", size=8.5, bold=True, color=BURGUNDY)
    ws["A5"].font = Font(name="Arial", size=8.5, bold=True, color=BURGUNDY)
    ws.merge_cells("B4:G4")
    ws.merge_cells("B5:G5")

    con = zotero_connection()
    if con is None:
        total, counts, items = 0, [], []
        ws["B4"] = f"No se encontró {ZOTERO_DB}"
    else:
        total, counts = get_zotero_counts(con)
        source_ids = [
            442,
            457,
            485,
            499,
            528,
            535,
            538,
            541,
            564,
            612,
            630,
            632,
            644,
            695,
            777,
            783,
        ]
        items = get_zotero_items(con, source_ids)
        con.close()

    ws["A7"] = "Registros bibliográficos actuales"
    ws["B7"] = total
    set_cell(
        ws["A7"],
        font=Font(name="Arial", size=9, bold=True, color=WHITE),
        fill=PatternFill("solid", fgColor=BURGUNDY),
        alignment=Alignment(wrap_text=True),
        border=ALL_THIN,
    )
    set_cell(
        ws["B7"],
        font=Font(name="Arial", size=11, bold=True, color=BURGUNDY),
        fill=PatternFill("solid", fgColor=PALE_RED),
        alignment=Alignment(horizontal="center"),
        border=ALL_THIN,
    )

    count_headers = ["Categoría Zotero", "Tag", "N.° de registros"]
    for idx, text in enumerate(count_headers, 1):
        set_cell(
            ws.cell(9, idx),
            value=text,
            font=Font(name="Arial", size=8.5, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=TEAL),
            alignment=Alignment(horizontal="center", wrap_text=True),
            border=ALL_THIN,
        )
    for r_idx, (label, tag, count) in enumerate(counts, 10):
        ws.cell(r_idx, 1).value = label
        ws.cell(r_idx, 2).value = tag
        ws.cell(r_idx, 3).value = count
        style_range(
            ws,
            f"A{r_idx}:C{r_idx}",
            font=Font(name="Arial", size=8, color=DARK),
            fill=PatternFill("solid", fgColor=WHITE if r_idx % 2 else "F5F7F8"),
            alignment=Alignment(vertical="center", wrap_text=True),
            border=ALL_THIN,
        )
    if counts:
        ws.conditional_formatting.add(
            f"C10:C{9 + len(counts)}",
            ColorScaleRule(
                start_type="min",
                start_color=PALE_GOLD,
                mid_type="percentile",
                mid_value=50,
                mid_color=PALE_TEAL,
                end_type="max",
                end_color=TEAL,
            ),
        )

    start = 10 + len(counts) + 2
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    set_cell(
        ws.cell(start, 1),
        value="FUENTES CLAVE Y FUNCIÓN EN LA MATRIZ",
        font=Font(name="Arial", size=10, bold=True, color=WHITE),
        fill=PatternFill("solid", fgColor=BURGUNDY_DARK),
        alignment=Alignment(horizontal="center"),
        border=ALL_THIN,
    )
    source_headers = [
        "Año",
        "Autores",
        "Título",
        "Temas / tipos Zotero",
        "Función en la matriz",
        "Item ID",
        "Enlace Zotero",
    ]
    for idx, text in enumerate(source_headers, 1):
        set_cell(
            ws.cell(start + 1, idx),
            value=text,
            font=Font(name="Arial", size=8.5, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=GREEN),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=ALL_THIN,
        )

    roles = {
        442: "Fundamenta el problema de ambientes térmicos desfavorables.",
        457: "Benchmark de ampacidad y validación del modelo.",
        485: "Línea base normativa de ampacidad en régimen permanente.",
        499: "Marco de rating térmico dinámico y variables operativas.",
        528: "Norma para carga cíclica y secado parcial del suelo.",
        535: "Evidencia reciente del impacto del secado del suelo.",
        538: "Benchmark multicapa, bedding/PAC y efecto de temperaturas de entorno.",
        541: "Concepto de cuello de botella térmico y localización crítica.",
        564: "Aporte sobre optimización de bedding y propiedades térmicas.",
        612: "Estado del arte y limitaciones metodológicas de PINN.",
        630: "Aplicación PINN a reconstrucción térmica 2D.",
        632: "PINN directo/inverso para conducción estacionaria.",
        644: "Panorama de evaluación térmica y current rating de cables.",
        695: "Mapa sistemático reciente de DTR para cables subterráneos.",
        777: "Fundamento y riesgos de Scientific ML/PINN.",
        783: "Medición, incertidumbre y factores de conductividad térmica del suelo.",
    }
    for r_idx, item in enumerate(items, start + 2):
        year = item["date"][:4] if item["date"] else "s. f."
        selected_tags = [
            tag
            for tag in item["tags"].split(",")
            if tag.startswith("tema-")
            or tag in {
                "tipo-estado-del-arte",
                "tipo-aporte",
                "tipo-aplicacion",
                "tipo-normativa",
                "tipo-guia-tecnica",
                "tipo-fuente-clasica",
                "thesis-core",
            }
        ]
        hyperlink = f"zotero://select/library/items/{item['key']}"
        values = [
            year,
            item["authors"],
            item["title"],
            "; ".join(selected_tags),
            roles.get(item["item_id"], "Sustento complementario."),
            item["item_id"],
            "Abrir en Zotero",
        ]
        for c_idx, value in enumerate(values, 1):
            set_cell(
                ws.cell(r_idx, c_idx),
                value=value,
                font=Font(name="Arial", size=8, color=DARK),
                fill=PatternFill("solid", fgColor=WHITE if r_idx % 2 else "F5F7F8"),
                alignment=Alignment(vertical="top", wrap_text=True),
                border=ALL_THIN,
            )
        ws.cell(r_idx, 7).hyperlink = hyperlink
        ws.cell(r_idx, 7).style = "Hyperlink"
        ws.cell(r_idx, 7).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[r_idx].height = 62

    widths = {"A": 10, "B": 28, "C": 44, "D": 33, "E": 34, "F": 10, "G": 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[start + 1].height = 34
    ws.freeze_panes = f"A{start + 2}"
    ws.print_title_rows = "1:5"
    ws.print_area = f"A1:G{start + 1 + len(items)}"
    configure_page(ws, landscape=True, paper="8", fit_height=3)
    return ws


def validate_workbook(path):
    wb = load_workbook(path, data_only=False)
    expected = {
        "Matriz UNI": (9, 6),
        "Matriz por actividades": (10, 10),
        "Operacionalización": (19, 9),
        "Sustento Zotero": (1, 7),
    }
    missing = [name for name in expected if name not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Hojas faltantes: {missing}")
    for name, (min_rows, min_cols) in expected.items():
        ws = wb[name]
        if ws.max_row < min_rows or ws.max_column < min_cols:
            raise RuntimeError(
                f"Hoja incompleta {name}: {ws.max_row}x{ws.max_column}"
            )
    required_phrases = [
        "heterogeneidad térmica",
        "temperatura máxima",
        "ampacidad",
        "modelo numérico 2d",
    ]
    text = " ".join(
        str(cell.value or "")
        for row in wb["Matriz UNI"].iter_rows()
        for cell in row
    ).lower()
    absent = [phrase for phrase in required_phrases if phrase not in text]
    if absent:
        raise RuntimeError(f"Contenido clave ausente: {absent}")


def main():
    MATRIX_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.properties.title = "Matriz de consistencia - heterogeneidad térmica en cables"
    wb.properties.subject = TITLE
    wb.properties.creator = AUTHORS
    wb.properties.description = (
        "Matriz elaborada a partir del proyecto, las guías UNI, el ejemplo de Excel "
        "y la base Zotero vigente."
    )
    build_official_matrix(wb)
    build_activity_matrix(wb)
    build_operationalization(wb)
    build_sources(wb)
    wb.active = 0
    wb.save(OUTPUT)
    validate_workbook(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
